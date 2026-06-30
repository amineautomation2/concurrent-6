import json
from math import ceil
from os import chdir
from os.path import dirname
from random import uniform
from re import sub
from time import sleep

import curl_cffi
import openpyxl

from utils import (
    clean_spreadsheet,
    delay,
    fetch_with_backoff,
    get_random_user_agent,
    get_xlsx_filepath,
    write_json,
)


def get_isin(fund_info: dict) -> str | None:
    has_isin = fund_info.get("ISIN")
    if has_isin:
        print("fundino ISIN")
        return has_isin

    type_code = fund_info.get("TypeCode")

    if type_code == "" or not type_code:
        return None
    cookies = {
        "ASP.NET_SessionId": "anocxi352i1cpdcexgg2mjrw",
    }

    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        "Origin": "https://digital-tools.feprecisionplus.com",
        "Sec-Fetch-Storage-Access": "none",
        "Sec-GPC": "1",
        "Connection": "keep-alive",
        "Referer": "https://digital-tools.feprecisionplus.com/",
        # 'Cookie': 'ASP.NET_SessionId=anocxi352i1cpdcexgg2mjrw',
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        # Requests doesn't support trailers
        # 'TE': 'trailers',
    }
    headers.update(get_random_user_agent())
    response = curl_cffi.get(
        f"https://digital.feprecisionplus.com/chelseafinancial/responsivecharting/gettableinstrumentdata/en-gb?jsonstring=%7B%22GrsProjectId%22:16000034,%22ToolId%22:19,%22LanguageCode%22:%22en-gb%22,%22ProjectName%22:%22chelseafinancial%22,%22PerformancePeriodEndType%22:3,%22DiscretePerformanceEndType%22:3,%22RiskAnalyticalPeriod%22:0,%22RiskBenchmark%22:0,%22RiskFreeRate%22:%22%22,%22CurrencyCode%22:%22GBP%22,%22CodeList%22:%5B%22{type_code}%22%5D,%22IsDisplayLatestAsAtDate%22:false%7D",
        cookies=cookies,
        headers=headers,
        impersonate="chrome",
    )

    if response.status_code != 200:
        print(f"{fund_info.get('Name')}: {response.status_code}")
    data = response.json()
    instrument = data.get("InstrumentInformations")
    if len(instrument) > 0:
        return instrument[0]["FundInfo"]["ISIN"]
    return None


def page_payload(page: int):
    payload = {
        "FilteringOptions": {
            "undefined": 0,
            "RangeId": "17200457",
            "Yield": {},
            "RangeName": "",
            "CategoryId": None,
            "Category2Id": None,
            "PriipProductCode": None,
            "DefaultCategoryId": None,
            "DefaultCategory2Id": None,
            "ForSaleIn": None,
            "ShowMainUnits": False,
            "MPCategoryCode": None,
        },
        "ProjectName": "chelseafinancial",
        "LanguageCode": "en-gb",
        "LanguageId": "1",
        "Theme": "chelsea",
        "SortingStyle": "1",
        "PageNo": page,
        "PageSize": 100,
        "OrderBy": "UnitName:init",
        "IsAscOrder": True,
        "OverrideDocumentCountryCode": None,
        "ToolId": "1",
        "PrefetchPages": 200,
        "PrefetchPageStart": page,
        "OverridenThemeName": "chelsea",
        "ForSaleIn": "",
        "ValidateFeResearchAccess": False,
        "HasFeResearchFullAccess": False,
        "EnableSedolSearch": "false",
        "GrsProjectId": "16000034",
        "ShowMainUnitExpansion": False,
        "UseCombinedOngoingChargeTER": False,
    }
    return json.dumps(payload)


def chelsea_total_pages() -> int:
    headers = {
        "Accept": "*/*",
        "Accept-Language": "en-GB,en;q=0.5",
        "Origin": "https://digital.feprecisionplus.com",
        "DNT": "1",
        "Connection": "keep-alive",
        "Referer": "https://digital.feprecisionplus.com/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    headers.update(get_random_user_agent())

    page = 1
    total_pages = 0
    payload_encoded = page_payload(page)
    url = f"https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetRowIdList?jsonString={payload_encoded}"

    res = fetch_with_backoff(
        url,
        headers=headers,
    )
    if res is not None:
        json_data = res.json()
        total_pages = ceil(json_data["TotalRows"] / 100)
        return total_pages
    return 0


def chelsea_runner() -> list[dict]:
    headers = {
        "Accept": "*/*",
        "Accept-Language": "en-GB,en;q=0.5",
        "Origin": "https://digital.feprecisionplus.com",
        "DNT": "1",
        "Connection": "keep-alive",
        "Referer": "https://digital.feprecisionplus.com/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    headers.update(get_random_user_agent())

    page = 1
    total_pages = 0

    has_next = True
    funds = []
    while has_next:
        payload_encoded = page_payload(page)
        url = f"https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetRowIdList?jsonString={payload_encoded}"

        # print(url)
        res = fetch_with_backoff(
            url,
            headers=headers,
        )
        # print(res)
        if res is not None:
            json_data = res.json()
            if page == 1:
                total_pages = ceil(json_data["TotalRows"] / 100)
                print(f"[###] Total Funds found: {json_data['TotalRows']} [###]")
            print(f"[#] Page {page} of {total_pages}")
            data = sub(r"\r\n\s", "", json_data["Units"])
            funds_per_page = json.loads(data)["DataList"]
            fund_base_url = "https://digital.feprecisionplus.com"
            fund_idx = 1
            for fund in funds_per_page:
                f = {}
                info = fund.get("FundInfo")
                print(f"[{fund_idx}/{len(funds_per_page)}]")
                if info:
                    name = info.get("Name")
                    if name:
                        f.update(dict(name=name))

                        # ws.cell(current_row_idx, 1, name)

                    has_url = info.get("FactsheetPdfLink")
                    if has_url:
                        fund_url = f"{fund_base_url}{has_url}"
                        f.update(dict(url=fund_url))

                    is_closed = info.get("ClosedToNewInvestors")
                    isin = info.get("ISIN")
                    f.update(dict(isin=isin))
                    code = info.get("TypeCode")
                    f.update(dict(type_code=code))
                    if is_closed:
                        f.update(dict(is_closed=is_closed))
                        # ws.cell(current_row_idx, 4, "Yes")
                    else:
                        f.update(dict(is_closed=is_closed))
                        # ws.cell(current_row_idx, 4, "No")
                funds.append(f)

                # ws.cell(current_row_idx, 1).value = name
                # ws.cell(current_row_idx, 2).value = isin
                # ws[f"D{current_row_idx}"] = "No" if fund["FundInfo"]["ClosedToNewInvestors"] else "Yes"
                fund_idx += 1
                delay(1, 2)
                if fund_idx == 5:
                    break
        has_next = False if page == total_pages else True
        page += 1
        break
    return funds


def write_data(funds: list):
    output_xlsx = get_xlsx_filepath("chelsea_financial.xlsx")
    clean_spreadsheet(output_xlsx)
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Funds"]
    current_row_idx = 2

    row_idx = 2
    for fund in funds:
        ws.cell(row_idx, 1, fund.get("name"))
        url = fund.get("url")
        c = ws.cell(row_idx, 3, url)
        c.hyperlink = url
        c.style = "Hyperlink"
        ws.cell(row_idx, 4, fund.get("is_closed"))
        isin = fund.get("isin")
        if isin is None:
            isin = get_isin(fund)

        ws.cell(row_idx, 2, isin)
        row_idx += 1
    wb.save(output_xlsx)
    wb.close()


def process_per_worker(id_w: int, max_w: int, total: int) -> list[dict]:
    headers = {
        "Accept": "*/*",
        "Accept-Language": "en-GB,en;q=0.5",
        "Origin": "https://digital.feprecisionplus.com",
        "DNT": "1",
        "Connection": "keep-alive",
        "Referer": "https://digital.feprecisionplus.com/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    headers.update(get_random_user_agent())

    pages = [x for x in range(1, total + 1)]
    pages = pages[id_w::max_w]
    funds = []
    for page in pages:
        payload_encoded = page_payload(page)
        url = f"https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetRowIdList?jsonString={payload_encoded}"

        # print(url)
        res = fetch_with_backoff(
            url,
            headers=headers,
        )
        if res is not None:
            json_data = res.json()
            if page == 1:
                print(f"[###] Total Funds found: {json_data['TotalRows']} [###]")
            print(f"[#] Page {page} of {pages}")
            data = sub(r"\r\n\s", "", json_data["Units"])
            funds_per_page = json.loads(data)["DataList"]
            fund_base_url = "https://digital.feprecisionplus.com"
            fund_idx = 1
            for fund in funds_per_page:
                f = {}
                info = fund.get("FundInfo")
                print(f"[{fund_idx}/{len(funds_per_page)}]")
                if info:
                    name = info.get("Name")
                    if name:
                        f.update(dict(name=name))

                        # ws.cell(current_row_idx, 1, name)

                    has_url = info.get("FactsheetPdfLink")
                    if has_url:
                        fund_url = f"{fund_base_url}{has_url}"
                        f.update(dict(url=fund_url))

                    is_closed = info.get("ClosedToNewInvestors")
                    isin = info.get("ISIN")
                    f.update(dict(isin=isin))
                    code = info.get("TypeCode")
                    f.update(dict(type_code=code))
                    if is_closed is not None:
                        f.update(dict(open=not is_closed))
                        # ws.cell(current_row_idx, 4, "Yes")
                        # ws.cell(current_row_idx, 4, "No")
                funds.append(f)

                # ws.cell(current_row_idx, 1).value = name
                # ws.cell(current_row_idx, 2).value = isin
                # ws[f"D{current_row_idx}"] = "No" if fund["FundInfo"]["ClosedToNewInvestors"] else "Yes"
                fund_idx += 1
                # if fund_idx == 3:
                #    break
                delay(1, 2)
    return funds
